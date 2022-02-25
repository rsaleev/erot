import os
from typing import Callable, List, Tuple, Union

from aiohttp import ClientSession
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet._read_only import ReadOnlyCell, ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet
from src.api.base.extractor import DocumentExtractor
from src.api.base.objects import Attribute, Object
from src.api.base.schema import DocumentSchemaResponse, HeaderAttribute
from src.api.exceptions import SheetValidationError, WorkbookNotFound

from src.api.transformers.excel import ExcelTransformer

import time


class ExcelExtractor(DocumentExtractor):

    __slots__ = ('workbook', 'worksheet', '_min_row', '_max_row', '_min_col',
                 '_max_col', '_document', '_header', '_sheetname', '_transformer', '_filename', '_cur_row')

    def __init__(self):
        self.workbook: Workbook
        self.worksheet: Union[Worksheet, ReadOnlyWorksheet]
        self._min_row = 1
        self._max_row = int(os.environ.get('SHEET_MAX_ROW', '1054000'))
        self._min_col = 1
        self._max_col = int(os.environ.get('SHEET_MAX_COL', '1000'))
        self._cur_row = 1
        self._document: Object
        self._header: Union[List[HeaderAttribute], None]
        self._transformer = ExcelTransformer()

    def read(self, filename: str):
        """
        read 

        Загрузка файла/рабочей книги

        Args:
            filename (str): имя файла/путь

        Raises:
            WorkbookNotFound: рабочая книга/файл не найден по указанному пути
        """
        # открытие книги/файла
        try:
            self.workbook = load_workbook(filename,
                                           data_only=True,read_only=True)
        except:
            raise WorkbookNotFound(f'Файл {filename} не найден')

    async def _send_validation_request(self,
                                       row: tuple) -> DocumentSchemaResponse:
        # генерация query params
        # headers=Заголовок1&headers=Заголовок2&...
        params = '&'.join([
            f'headers={cell.value}' for cell in row
            if cell.value and cell.value != '' and cell.value is not None
        ])
        if not params:
            return DocumentSchemaResponse(error='Пустой заголовок', data=None)
        # отправка запроса
        async with ClientSession() as session:
            async with session.get(url=os.environ['VALIDATION_URL'],
                                   params=params) as r:
                response = DocumentSchemaResponse(**await r.json())
                return response

    def _set_document(self, response: DocumentSchemaResponse):
        # инициализация структуры
        if not response.data:
            return
        self._document = Object(response.data.name)
        [self._document.add_field(column) for column in response.data.columns]

    def _set_header(self, response: DocumentSchemaResponse):
        # установка заголовка документа
        if response.data and response.data.header:
            self._header = response.data.header

    def _set_dimensions(self, row):
        """
        Определение размерности документа

        Args:
            row ([type]): [description]
        """
        # минимальная строка -> индекс строки с заголовком + 1 строка вниз
        self._min_row = row[0].row + 1
        self._max_col = max(col.document.index
                            for col in self._document.attributes)
        for row in self.worksheet.iter_rows(min_row=self._min_row,
                                            min_col=1,
                                            max_col=1,
                                            max_row=self._max_row):
            try:
                self._max_row = row[0].row
            except AttributeError:
                break

    async def _validate(self):
        """
        Проверка первых 10 строк на соответствие схемам заголовков документов

        Raises:
            SheetNotValid: при окончании итерации и отсутствии совпадения (включая отсутствующие обязательные заголовки столбцов)
        """
        error = 'Не удалось соотнести структуру документа и схему'
        valid = False
        for sheet in self.workbook.sheetnames:
            ws = self.workbook[sheet]
            if not isinstance(ws, (ReadOnlyWorksheet, Worksheet)):
                continue
            for row in ws.iter_rows(min_row=1,
                                    max_row=10,
                                    max_col=self._max_col):
                response = await self._send_validation_request(row)
                if response.error:
                    error = response.error
                    continue
                elif response.data:
                    valid = True
                    # назначение рабочего листа
                    self.worksheet = ws
                    # назначение структуры документа
                    self._set_document(response)
                    self._set_header(response)
                    self._set_dimensions(row)
                    break
        if not valid:
            raise SheetValidationError(error)

    def _set_value(self, attribute: Attribute, row: Tuple[ReadOnlyCell]):
        # определение значений по индексу, co смещением на 1
        attribute.value = next(value for idx, value in enumerate(row, start=1) if idx == attribute.document.index)

    async def extract(self, filename:str, callback):
        self.read(filename)
        await self._validate()            
        """
        extract 

        [extended_summary]

        Raises:
            StopIteration
        Returns:
            Object: [description]
        """
        for r in self.worksheet.iter_rows(min_row=self._min_row,
                                     max_row=self._max_row,
                                     min_col=self._min_col,
                                     max_col=self._max_col,
                                     values_only=True):
            t1 = time.time()
            document = self._document
            [self._set_value(attr, r) for attr in document.attributes]
            self._transformer.transform(document)
            t2 = time.time()
            await callback(document)
            t3 = time.time()
            print(t2-t1, t3-t2)
        self.workbook.close()


